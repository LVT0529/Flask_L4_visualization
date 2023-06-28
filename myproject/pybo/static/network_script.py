#selenium, openpyxl, pillow 모듈 설치 필요
#pip install selenium
#pip install openpyxl
#pip install Pillow

## 웹 크롤링 관련 라이브러리
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import operator
import time
import os
import re
import sys

## 엑셀 관련 라이브러리
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import openpyxl
import datetime

import pytesseract

#pytesseract.pytesseract.tesseract_cmd = r'C:\Users\victory\AppData\Local\Tesseract-OCR\tesseract'

# Headless Web 설정
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument("disable-gpu")
options.add_argument('window-size=1920x1080')
options.add_argument("lang=ko_KR")
driver = webdriver.Chrome(options=options, executable_path = os.getcwd() + '\pybo\static\chromedriver')


## 웹 드라이브 실행
#driver = webdriver.Chrome()


#############################그래프 캡처 함수#############################
def capture(startdate, enddate, graph_list):
    #Total Template 선택
    select = Select(driver.find_element_by_id("graph_template_id"))
    select.select_by_value('70')

    #시작 시간 설정
    driver.find_element_by_id("date1").clear()
    date = driver.find_element_by_id("date1")
    date.send_keys(startdate)

    #종료 시간 설정
    driver.find_element_by_id("date2").clear()
    date = driver.find_element_by_id("date2")
    date.send_keys(enddate)
    driver.find_element_by_name('button_refresh_x').click()


    #검색
    '''
    elem = driver.find_element_by_name("filter")
    elem.send_keys(search)
    elem.send_keys(Keys.RETURN)
    '''



    list = []
    count = 0

    date = re.split(" |-", startdate)
    date = date[0]+"-"+date[1]

    for tem in range (1,3):
        for num in range(0,len(graph_list)):
            for i in range(2,35):
                for j in range (1,4):
                    try:
                        #select = Select(driver.find_element_by_id("host_id"))
                        #select.select_by_value(host)
                        #path = '//*[@id="main"]/table[2]/tbody/tr/td/table/tbody/tr[' + str(i) + ']/td['+ str(j) +']/table/tbody/tr/td[2]/a[1]/img'
                        #driver.find_element_by_xpath(path).click()
                        path = '//*[@id="main"]/table[2]/tbody/tr/td/table/tbody/tr[' + str(i) + ']/td['+ str(j) +']/table/tbody/tr/'
                        graph_id = driver.find_element_by_xpath(path + 'td[1]/div/a/img').get_attribute("id")

                        if(graph_id in graph_list[num]):
                            driver.find_element_by_xpath(path + 'td[2]/a[1]/img').click()
                            element1 = driver.find_element_by_class_name('graphimage')
                            element_png = element1.screenshot_as_png

                            graph_name = element1.get_attribute("alt")
                            graph_name = graph_name.split("[")

                            list.append("%s  %s.png" %(date, graph_name[0]))
                            with open("%s  %s.png" %(date, graph_name[0]),"wb") as file:
                                file.write(element_png)
                            print(graph_name[0] + "  Saveing .....")
                            driver.back()
                    except:
                        break
                        time.sleep(1)
        # Global Connection Template 선택
        select = Select(driver.find_element_by_id("graph_template_id"))
        select.select_by_value('38')

    return list

#############################그래프 이미지 엑셀#############################
def Capture_Excel(list_before, list_after):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws = wb['Sheet']
    ws.title = 'Cacti 캡쳐'

    num = 1
    for i in list_before:
        img = Image(i)
        ws.add_image(img, 'A%d' %(num))
        num += 12

    num = 1
    for i in list_after:
        img = Image(i)
        ws.add_image(img, 'J%d' %(num))
        num += 12


    nowdate = datetime.datetime.now()

    print("Make Capture Excel....... ")
    wb.save(nowdate.strftime("%Y-%m-%d") + ' Montly Report.xlsx')

#############################VLAN 정렬 함수#############################
def vlan_list(ad_id, ad_pw):
    #driver = webdriver.Chrome()
    #driver.maximize_window()
    driver.get("http://114.200.198.228/phpipam/")

    elem = driver.find_element_by_name('ipamusername')
    elem.send_keys(ad_id)
    elem = driver.find_element_by_name("ipampassword")
    elem.send_keys(ad_pw)
    elem.send_keys(Keys.RETURN)

    time.sleep(3)

    driver.find_element_by_xpath('//ul[@class="nav navbar-nav icons"]/li[1]/a').click()

    path = "table sorted vlans table-condensed table-top bdt"
    table = driver.find_element_by_xpath('//table[@class="' + path + '"]/tbody')
    tr = table.find_elements_by_tag_name("tr")

    ## Page 수 All로 선택
    select = Select(driver.find_element_by_id("page-rows-form"))
    select.select_by_value('1000000')

    list = ["even change", "odd change", "odd"]

    temp = []
    vlan = []
    ip_list = []
    floor = []
    sort_list = []

    for count in range(1, len(tr)+1):
        class_name = driver.find_element_by_xpath('//table[@class="' + path + '"]/tbody/tr['+ str(count) + ']').get_attribute("class")

        if(class_name in list):
            temp = driver.find_element_by_xpath('//table[@class="' + path + '"]/tbody/tr['+ str(count) + ']').text.split(" ")
            if(len(temp) >= 3):
                vlan.append(temp[0])
                ip_list.append(temp[2])
                floor.append(temp[3])
            else:
                vlan.append(vlan[len(vlan)-1])
                ip_list.append(temp[0])
                floor.append(temp[1])

                temp = [0,0,0,0]
                temp[0] = vlan[len(vlan)-1]
                temp[1] = vlan[len(vlan)-1]
                temp[2] = ip_list[len(ip_list)-1]
                temp[3] = floor[len(floor)-1]

            sort_list.append(temp)

    sort_list.sort(key = operator.itemgetter(3))
    driver.quit()

    return sort_list

#############################VLAN 정렬 엑셀#############################
def Vlan_Excel(sort_list):

    nowdate = datetime.datetime.now()
    wb = openpyxl.load_workbook(nowdate.strftime("%Y-%m-%d") + ' Montly Report.xlsx')

    sheet = wb.create_sheet('Vlan 정리')

    count = 0
    for i in sort_list:
        sheet['A' +str(count + 1)] = sort_list[count][1]
        sheet['B' +str(count + 1)] = sort_list[count][2]
        sheet['C' +str(count + 1)] = sort_list[count][3]

        count += 1

    print("Make Vlan Excel....... ")
    wb.save(nowdate.strftime("%Y-%m-%d") + ' Montly Report.xlsx')

#############################Max traffic, tax 엑셀#############################
def Max_sort_Excel(files):
    nowdate = datetime.datetime.now()
    wb = openpyxl.load_workbook(nowdate.strftime("%Y-%m-%d") + ' Montly Report.xlsx')

    sheet = wb.create_sheet(nowdate.strftime("%Y-%m"))

    for i in files:
        if(i.filename == 'watch_max_all_d_all.xlsx'):
            watch_max_all_d_all = i
        elif(i.filename == 'broad_max_all_d_all.xlsx'):
            broad_max_all_d_all = i
        elif(i.filename == 'watch_max_bps_d_all_m.xlsx'):
            watch_max_bps_d_all_m = i
        elif(i.filename == 'watch_max_bps_d_all_p.xlsx'):
            watch_max_bps_d_all_p = i


    ## 최고 동시시청
    watch = openpyxl.load_workbook(watch_max_all_d_all)
    watch = watch.active

    ## 최고 동시방송 갯수
    broad = openpyxl.load_workbook(broad_max_all_d_all)
    broad = broad.active

    ## 화질별 PC 이용자 수
    pc = openpyxl.load_workbook(watch_max_bps_d_all_p)
    pc = pc.active

    ## 화질별 모바일 이용자 수
    mob = openpyxl.load_workbook(watch_max_bps_d_all_m)
    mob = mob.active

    ascii = 65
    #################### 최대 시청자 수 ####################
    sheet['C1'] = "시청수"
    for day in range(0,31):
        if(watch['C'+ str(36-day)].value == None):
            continue
        else:
            sheet[chr(ascii) + str(day+2)] = watch['B'+ str(36-day)].value  ## 날짜
            sheet[chr(ascii+2) + str(day+2)] = watch['C'+ str(36-day)].value  ## 시청자 수

    ascii += 3

    #################### 최대 방송 갯수 ####################
    sheet['D1'] = "방송개수"
    for day in range(0,31):
        if(broad['C'+ str(36-day)].value == None):
            continue
        else:
            sheet[chr(ascii) + str(day+2)] = broad['C'+ str(36-day)].value  ## 방송개수

    ascii +=2

    #################### 최대 화질별 PC 이용자 수 ####################
    sheet['F1'] = "전체"
    sheet['G1'] = "일반화질"
    sheet['H1'] = "고화질"
    sheet['I1'] = "원본화질"
    for day in range(0,31):
        if(pc['C'+ str(36-day)].value == None):
            continue
        else:
            sheet[chr(ascii) + str(day+2)] = pc['C'+ str(36-day)].value  ## 전체
            sheet[chr(ascii+1) + str(day+2)] = pc['D'+ str(36-day)].value  ## 일반화질
            sheet[chr(ascii+2) + str(day+2)] = pc['E'+ str(36-day)].value  ## 고화질
            sheet[chr(ascii+3) + str(day+2)] = pc['F'+ str(36-day)].value  ## 원본화질

    ascii +=5
    #################### 최대 화질별 모바일 이용자 수 ####################
    sheet['K1'] = "전체"
    sheet['L1'] = "앱-라디오모드"
    sheet['M1'] = "앱-저화질"
    sheet['N1'] = "앱-일반"
    sheet['O1'] = "앱-고화질"
    sheet['P1'] = "앱-최고화질(원본)"
    sheet['Q1'] = "웹-저화질"
    sheet['R1'] = "PC+앱 최고화질 시청자수"
    for day in range(0,31):
        if(mob['C'+ str(36-day)].value == None):
            continue
        else:
            sheet[chr(ascii) + str(day+2)] = mob['C'+ str(36-day)].value  ## 전체
            sheet[chr(ascii+1) + str(day+2)] = mob['D'+ str(36-day)].value  ## 앱-라디오
            sheet[chr(ascii+2) + str(day+2)] = mob['E'+ str(36-day)].value  ## 앱-저화질
            sheet[chr(ascii+3) + str(day+2)] = mob['F'+ str(36-day)].value  ## 앱-일반
            sheet[chr(ascii+4) + str(day+2)] = mob['G'+ str(36-day)].value  ## 앱-고화질
            sheet[chr(ascii+5) + str(day+2)] = mob['H'+ str(36-day)].value  ## 앱-최고화질
            sheet[chr(ascii+6) + str(day+2)] = mob['I'+ str(36-day)].value  ## 웹-저화질
            sheet[chr(ascii+7) + str(day+2)] = (mob['H'+ str(36-day)].value + pc['F'+ str(36-day)].value) ## PC+앱 최고화질 시청자수



    data = wb.create_sheet(nowdate.strftime("%Y-%m") + ' 데이터')
    ascii = 65
    #################### 요약 취합 본 ####################
    data['A1'] = "날짜"
    data['C1'] = "일 최고 동시 시청"
    data['D1'] = "원본화질(PC)"
    data['E1'] = "앱-최고화질(원본)"
    data['F1'] = "PC+앱 최고화질 시청자수"
    data['G1'] = "일 최고 동시 방송"
    for day in range(2,33):
        if(mob['C'+ str(37-day)].value == None):
            continue
        else:
            data[chr(ascii) + str(day)] = sheet['A'+ str(day)].value  ## 날짜
            data[chr(ascii+2) + str(day)] = sheet['C'+ str(day)].value  ## 일 최고 동시 시청
            data[chr(ascii+3) + str(day)] = sheet['I'+ str(day)].value  ## 원본화질(PC)
            data[chr(ascii+4) + str(day)] = sheet['P'+ str(day)].value  ## 앱-최고화질(원본)
            data[chr(ascii+5) + str(day)] = sheet['R'+ str(day)].value  ## PC+앱 최고화질 시청자수
            data[chr(ascii+6) + str(day)] = sheet['D'+ str(day)].value  ## 일 최고 동시 방송


    print("Make Max_sort Excel....... ")
    wb.save(nowdate.strftime("%Y-%m-%d") + ' Montly Report.xlsx')


#############################트래픽 수치# #############################
def Image_text_extract (list_before, list_after):

    for num in range(0, len(list_before)):
        list = pytesseract.image_to_string(list_before[num], lang = 'kor+eng', config = '-c preserve_interword_spaces=1  --psm 4')
        temp = re.split(" |\n", list)

        inbound = temp[len(temp)-13] + temp[len(temp)-12]
        outbound = temp[len(temp)-3] + temp[len(temp)-2]

        print(inbound)
        print(outbound)

    for num in range(0, len(list_after)):
        list = pytesseract.image_to_string(list_after[num], lang = 'kor+eng', config = '-c preserve_interword_spaces=1  --psm 4')
        temp = re.split(" |\n", list)

        inbound = temp[len(temp)-13] + temp[len(temp)-12]
        outbound = temp[len(temp)-3] + temp[len(temp)-2]

        print(inbound)
        print(outbound)

def network_report(ad_id, ad_pw, startdate_before, enddate_before, startdate_after, enddate_after, files):

    sort_list = []

    driver.maximize_window()
    driver.get("http://211.49.171.30/cacti/index.php")

    username = "admin"
    password = "7msanwk"

    time.sleep(1)
    elem = driver.find_element_by_name('login_username')
    elem.send_keys(username)
    time.sleep(1)
    elem = driver.find_element_by_name("login_password")
    elem.send_keys(password)
    elem.send_keys(Keys.RETURN)


    now_time = time.strftime('%y-%m-%d')
    ## Graph_Capture 폴더 생성
    if not os.path.isdir("Graph_Capture " + now_time):
       os.makedirs("Graph_Capture " + now_time)
       print("Make directory 'Graph_Capture'")
    os.chdir(os.getcwd() + "\Graph_Capture " + now_time)


    # Graph 이미지  클릭
    driver.find_element_by_xpath('//*[@id="tabs"]/a[2]/img').click()



    list_before = []
    list_after = []

    graph_list = ['graph_37008', 'graph_22487', 'graph_14361', 'graph_26661', 'graph_23318',
                  'graph_29243', 'graph_14907', 'graph_14924', 'graph_1008', 'graph_1485',
                  'graph_11347', 'graph_11396', 'graph_14928', 'graph_14945', 'graph_27342',
                  'graph_27350', 'graph_27836', 'graph_27849']  #graph_22592부터 미러서버 트래픽

    # Cacti 이미지 생성
    list_before = capture(startdate_before, enddate_before, graph_list)
    list_after = capture(startdate_after, enddate_after, graph_list)

    # Cacti 그래프이미지 엑셀 생성
    Capture_Excel(list_before, list_after)

    # 이미지 텍스트 추출
    #Image_text_extract(list_before, list_after)

    # Vlan 생성
    sort_list = vlan_list(ad_id, ad_pw)

    # Vlan 엑셀 생성
    Vlan_Excel(sort_list)

    # Vlan 엑셀 생성
    Max_sort_Excel(files)

    driver.quit()
